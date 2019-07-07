import openpyxl
path='E:\\python_my_practice\\filehandling\\stud.xlsx'
'''
print('read the value from excel..')
sheet=openpyxl.load_workbook(path)
sheet_obj=sheet.active
cell_obj=sheet_obj.cell(1,1)
print(cell_obj.value)

sheet = openpyxl.load_workbook(path)
sheet_obj=sheet.active
# print(cell_obj.value)
# print('total no of rows..')
# print(sheet_obj.max_column)

# print('total no of columns...')
# print(sheet_obj.max_row)


m_row=sheet_obj.max_row
for col in range(1,m_obj+1):
    cell_obj=sheet_obj.cell(row=col,column=1)
    print(cell_obj.value)
'''

'''
print('print the perticular row value..')
sheet=openpyxl.load_workbook(path)
sheet_obj=sheet.active

m_col=sheet_obj.max_column
for row in range(1,m_col+1):
    cell_obj=sheet_obj.cell(row=1,column=row)
    print(cell_obj.value,end=' ')
'''


