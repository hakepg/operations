import openpyxl
'''
sheet=openpyxl.Workbook()

sheet_ob=sheet.active
# sheet_ob.title #print title to the sheet
# print('sheet title:',sheet_ob)

sheet_ob.title='pgh' # rename the title of sheet
print('sheet rename title:',sheet_ob)
'''
'''
print('prog to write to an excel file')
path='E:\\python_my_practice\\filehandling\\emp.xlsx'
sheet=openpyxl.Workbook()

sh=sheet.active

c1=sh.cell(1,1)
c1.value='name'
c2=sh.cell(1,2)
c2.value='age'
c3=sh.cell(1,3)
c3.value='salary'
c4=sh.cell(1,4)
c4.value='address'
c5=sh.cell(2,1)
c5.value='pratima'
sheet.save(path)
'''




'''
print('prog to add sheet into workbook..')
wb=openpyxl.Workbook()

sheet=wb.active

wb.create_sheet(index=1,title='demo sheet')

wb.save('E:\\python_my_practice\\filehandling\\emp.xlsx')
'''











