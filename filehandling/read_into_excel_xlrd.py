import openpyxl
import xlrd
path='E:\\python_my_practice\\filehandling\\stud.xlsx' #give the path of file storeage location

'''
sheet=xlrd.open_workbook(path)
index=sheet.sheet_by_index(0)

print(index.cell_value(1,1))
print(index.cell_value(1,2))
print(index.cell_value(1,3))
print(index.cell_value(1,4))
'''


'''
print('Extract value by row wise and column wise')

sheet=xlrd.open_workbook(path) # open the file
wb=sheet.sheet_by_index(0) # 1st give the index wise
sheet1=wb.cell_value(0,0) #call the cell by row and column wise

print('extract by col wise is {}'.format(wb.ncols))
print('extract by row wise is {}'.format(wb.nrows))
'''

print('extracting all column names')

sheet=xlrd.open_workbook(path)
wb=sheet.sheet_by_index(0)
sh1=wb.cell_value(0,0)

for exl in range(wb.ncols):
    print(wb.cell_value(1,exl))


































# wb_obj=openpyxl.load_workbook(path)
#
# sheet_obj=wb_obj.active
# cell_obj1=sheet_obj.cell(row=1, column=1)
# cell_obj2=sheet_obj.cell(row=2, column=1)
# cell_obj3=sheet_obj.cell(row=1, column=2)
#
# print(cell_obj1.value)
# print(cell_obj2.value)
# print(cell_obj3.value)

