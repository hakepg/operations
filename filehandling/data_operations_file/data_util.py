import pymysql
import openpyxl


connection=None
def get_db_conn():
    global connection

    if connection == None:
        connection = pymysql.connect('localhost','root','agh123','')
    return connection


file_path = 'E:\\python_my_practice\\filehandling\\data_operations_file\\'
def get_excel_sheet():
    wb=openpyxl.load_workbook(file_path+'pratima.xlsx') # load the workbook for adding the data
    sheet=wb['Empinfo'] #write data into that sheet
    return sheet,wb  #

   # wb=openpyxl.Workbook() #create workbook
   # wb.create_sheet('Empinfo')  # create sheet of name empinfo
   # wb.save(file_path+'pratima.xlsx')  #save that sheet in ur workbook using the path of storing location and name of excelsheet or workbook


get_excel_sheet()


# def get_text_file():
#     return open(file_path+'empinfo.txt','+a')
# get_text_file()